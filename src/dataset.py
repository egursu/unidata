import os
import csv
import json
from typing import Self
from io import TextIOWrapper, BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, numbers
from .database import Database
from .ssh import SSH
from .constants import XLSX_MAX_ROWS
from .utils import iter_lowered, yield_list, file_extension, to_list, iter_in_str


class Dataset:
    data: list = []
    columns: list = []
    extra_data: list = []
    column_prefix: str = "Col"

    def __init__(self, data=None, columns: list|tuple|str=None, **kwargs) -> None:
        if data:
            # from records
            if isinstance(data, (list, tuple, set)):
                self.from_records(data, columns)
            # from string as filename (xls or csv)
            elif isinstance(data, str) and iter_in_str(("xls", "csv", ), file_extension(data)): 
                self.from_excel(data, **kwargs) if 'xls' in file_extension(data) else self.from_csv(data, **kwargs)
            # from database query
            elif (isinstance(data, str) and kwargs.get("sql_text")) or iter_in_str(("Connection", "Database", ), type(data).__name__):
                self.from_sql(data, **kwargs)
            # from string splitted string by , or int/float value
            elif isinstance(data, (str, int, float)):
                    self.from_records(self.as_matrix(to_list(data)), to_list(columns))
            # from dictionary
            elif isinstance(data, dict):
                self.from_dict(data)
            # from dataset
            elif isinstance(data, Dataset):
                self.from_dataset(data)
            # from openpyxl Worksheet
            elif "Worksheet" in type(data).__name__:
                self.from_worksheet(data, **kwargs)
            else:
                self.data = self.as_matrix(data)
        self.extra_data = kwargs.get("extra_data", self.extra_data)
        if self.extra_data:      
            self.extra_data = self.as_matrix(self.extra_data)
        self.column_prefix = kwargs.get("column_prefix", self.column_prefix)

    def copy(self):
        return Dataset(self.data, self.columns)
    
    def from_records(self, data: list|tuple|set, columns: list|tuple|str=None) -> Self:
        if isinstance(data, set):
            data = list(data)
        if data and isinstance(data[0], dict):
            self.data = [list(row.values()) for row in data]
            self.columns  = list(data[0].keys())
        else:
            self.data = self.as_matrix(data)
            self.columns = to_list(columns)
        return self

    def from_dict(self, data: dict) -> Self:
        self.data = self.as_matrix(list(data.values()))
        self.columns = list(data.keys())
        return self

    def from_dataset(self, data: Self) -> Self:
        self.data = list(data.data)
        self.columns = list(data.columns)
        return self

    def from_sql(self, database, sql_text: str, sql_params=None, as_list: bool=False) -> Self:
        def itercursor(cursor, batch_size=100000):
            while True:
                rows = cursor.fetchmany(batch_size)
                if not rows:
                    break
                yield rows
        is_url = isinstance(database, str)
        if is_url:
            database = Database(database)
        cursor = database.connection.cursor()
        cursor.execute(sql_text, sql_params)
        self.columns = [desc[0] for desc in cursor.description]
        # data = cursor.fetchall()
        data = []
        for rows in itercursor(cursor):
            data.extend([list(row) for row in rows])
        self.data = [list(row) for row in data] if as_list else data
        del cursor
        if is_url:
            database.close()
        return self

    def from_csv(self, file_name: str, delimiter: str=',', quotechar: str='"', quoting=csv.QUOTE_MINIMAL, column_index: int|bool=1, encoding: str='utf-8', ssh=None) -> Self:
        is_ssh_instance = not isinstance(ssh, str)
        if ssh:
            if isinstance(ssh, str):
                ssh = SSH(ssh)
            csv_file = TextIOWrapper(ssh.sftp.file(os.path.join(ssh.path, file_name)), encoding=encoding)
        else:
            csv_file = open(file_name, encoding=encoding)
        with csv_file:
            csv_reader = csv.reader(csv_file, delimiter=delimiter, quotechar=quotechar, quoting=quoting)
            data = list(csv_reader)
            self.data = data[int(column_index):]
            if column_index:
                self.columns = data[column_index-1]
        if ssh and not is_ssh_instance:
            ssh.close()
        return self
    
    def from_worksheet(self, worksheet, column_index: int|bool=1, extra_data:bool=True, empty_cols:bool=False) -> Self:
        if not empty_cols: # remove empty cols
            for i in sorted((j for j, cell in enumerate(list(worksheet.rows)[column_index-1], start=1) if not cell.value), reverse=True):
                worksheet.delete_cols(i)
        data = list(worksheet.iter_rows(1, worksheet.max_row))
        self.data = [[cell.value for cell in row] for row in data[int(column_index):]]
        if column_index:
            self.columns = [cell.value or None for cell in data[int(column_index)-1]]
        if extra_data and int(column_index) > 1:
            self.extra_data = [[cell.value for cell in row] for row in data[:int(column_index)-1]]
        return self
    
    def from_excel(self, file_name: str, sheet_name: str=None, column_index: int|bool=1, extra_data:bool=True, empty_cols:bool=False) -> Self:
        workbook = load_workbook(file_name)
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        self.from_worksheet(worksheet, column_index, extra_data, empty_cols)
        workbook.close()
        return self

    def to_csv(self, file_name: str, delimiter: str=',', quotechar: str='"', quoting=csv.QUOTE_MINIMAL, header: bool=True, encoding: str='utf-8') -> None:
        with open(file_name, mode='w', newline='', encoding=encoding) as csv_file:
            csv_writer = csv.writer(csv_file, delimiter=delimiter, quotechar=quotechar, quoting=quoting, lineterminator='\n')
            if header:
                csv_writer.writerow(self.columns)
            csv_writer.writerows(self.data)

    def to_excel(self, file_name: str, sheet_name: str=None, formatted: bool=True, stream: bool=False) -> str|BytesIO:
        return dataset_to_excel(file_name, self, sheet_names=sheet_name, formatted=formatted, stream=stream)

    def to_dict(self) -> list[dict]:
        return [dict(zip(self.columns, row)) for row in self.data]

    def to_json(self, file_name: str=None) -> dict:
        json_data = json.dumps(self.to_dict(), indent = 4)
        if file_name:
            with open(file_name, "w") as file: 
                json.dump(json_data, file)
        return json_data
    
    def to_sql(self, database, table: str, auto_commit: bool=True) -> None:
        is_url = isinstance(database, str)
        if is_url:
            database = Database(database)
        cursor = database.connection.cursor() if is_url else database.cursor
        bind_columns = ",".join(self.columns)
        bind_params = ",".join([database.placeholder.format(column) for column in self.columns])
        if hasattr(cursor, 'fast_executemany'):
                cursor.fast_executemany = True
        cursor.executemany(f"INSERT INTO {table} ({bind_columns}) VALUES ({bind_params})", self.data) 
        if auto_commit:
            cursor.connection.commit()
        if is_url: 
            del cursor
            database.close()

    @property
    def columns_lowered(self) -> list:
        return iter_lowered(self.columns)
    
    @property
    def is_empty(self) -> list:
        return len(self.data[0]) == 0 if self.data else True
    
    @staticmethod
    def as_matrix(data) -> list:
        matrix = []
        if not data:
            return [[]]
        elif data and not isinstance(data[0], (list, tuple, dict, set)):
            matrix.append(data)
            return matrix
        return list(data)
    
    def columns_index(self, data) -> list:
        result = []
        for value in to_list(data, slice_stop=len(self.columns)):
            try:
                index = int(value)
            except (TypeError, ValueError):
                if not value.lower() in self.columns_lowered:
                    raise ValueError(f"Column '{value}' is not in Dataset columns")
                index = self.columns_lowered.index(value.lower())
            result.append(index)
        return result

    @staticmethod
    def max_value_len(data: list|tuple) -> list:
        return [len(max([str(row[i]) for row in data], key=len)) for i in range(len(data[0]))]
    
    def __getattribute__(self, name: str):
        if name == "columns":
            columns = self.__dict__.get(name)
            if (not columns) and self.data:
                self.columns = [f"{self.column_prefix}{i+1}" for i in range(len(self.data[0]))]
            if not isinstance(columns, list):
                self.columns = to_list(columns)
        return super().__getattribute__(name)

    def __getitem__(self, item) -> Self:
        item = to_list(item, slice_stop=len(self.columns))
        if item:
            # return list([data[idx] for idx in self.columns_index(item)] for data in self.data)
            data = list([data[idx] for idx in self.columns_index(item)] for data in self.data)
            columns = list(self.columns[idx] for idx in self.columns_index(item))
            return Dataset(data, columns)
        else:
            raise TypeError("Invalid Argument Type")
    
    def __setitem__(self, key, value) -> None:
        key = to_list(key, slice_stop=len(self.columns))
        value = self.as_matrix(to_list(value, slice_stop=len(self.columns)))
        for j, row in enumerate(self.data):
            for i, idx in enumerate(self.columns_index(key)):
                row[idx] = value[i][j]
    
    def values(self, columns=None) -> list:
        columns = self.columns_index(columns or self.columns)
        if columns:
            return list([data[idx] for data in self.data] for idx in columns)
        else:
            raise TypeError("Invalid Argument Type")

    def rename_columns(self, columns) -> Self:
        if isinstance(columns, dict):
            columns = {str(k).lower(): v for k, v in columns.items()}.get
            self.columns = list(map(columns, self.columns_lowered, self.columns))
        else:
            columns = to_list(columns)
            if len(self.columns) != len(columns):
                raise ValueError(f'Different length of fields (source: {len(self.columns)}, target: {len(columns)}).')
            self.columns = columns
        return self
    
    def remove(self, columns) -> Self:
        columns = self.columns_index(columns)
        columns.sort()
        for i, index in enumerate(columns):
            self.columns.pop(index - i)
        for item in self.data:
            for i, index in enumerate(columns):
                item.pop(index - i)
        return self

    def append_default_values(self, data: dict):
        data = {k: v for k, v in data.items() if k not in self.columns}
        self.columns.extend(data.keys())
        for item in self.data:
            item.extend(data.values())

    def auto_increment(self, columns, start: int=1):
        columns = self.columns_index(columns)
        for seq, value in enumerate(self.data, start=start):
            for idx in columns:
                value[idx] = seq

    def left_join(self, dataset, columns_by: str) -> Self:
        columns_by = columns_by.split("=")
        column_source = columns_by[0]
        column_target = column_source if len(columns_by) == 1 else columns_by[1]
        error_msg = "Columns {} not in dataset."
        try:
            source_index = self.columns_lowered.index(column_source.lower())
        except ValueError:
            raise ValueError(error_msg.format(column_source))
        try:
            target_index = iter_lowered(dataset.columns).index(column_target.lower())
        except ValueError:
            raise ValueError(error_msg.format(column_target))
        # append do data
        for row in self.data:
            target_row  = next((item for item in dataset.data if item[target_index]==row[source_index]), len(dataset.columns)*[None])
            row.extend([item for i, item in enumerate(target_row) if i != target_index])
        # rename duplicates columns
        column_target = [col for i, col in enumerate(dataset.columns) if i != target_index]
        for i, column in enumerate(column_target):
            if column.lower() in self.columns_lowered:
                cnt = 1
                while True:
                    column = f"{column_target[i]}_{cnt}"
                    if column.lower() not in self.columns_lowered:
                        break
                    cnt += 1
            self.columns.append(column)
        return self
    
    def query(self) -> Self:
        return self
    
    def union(self, data: list|tuple|Self) -> Self:
        data = data.data if isinstance(data, Dataset) else self.as_matrix(data)
        source_len = len(self.data[0])
        target_len = len(data[0])
        if source_len != target_len:
            raise ValueError(f'Different length of data (source: {source_len}, target: {target_len}).')
        self.data.extend(data)
        return self
    
    def unique(self, columns=None) -> Self:
        columns = self.columns_index(columns or self.columns)
        data = [list(item) for item in set(tuple(data[idx] for idx in columns) for data in self.data)]
        # return data
        columns = list(self.columns[idx] for idx in columns)
        return Dataset(data, columns)
        # return [list(item) for item in set(tuple(item) for item in self.values(columns))]

    def convert(self, to_type, columns=None) -> Self:
        columns = self.columns_index(columns or self.columns)
        for data in self.data:
            for idx in columns:
                data[idx] = to_type(data[idx])
        return self

    def __str__(self) -> str:
        if self.columns:
            if len(self.columns) != len(self.data[0]):
                raise ValueError(f'Different length of columns({len(self.columns)}) and data({len(self.data[0])}).')
            SHOW_ROWS = 3
            SHOW_COLS = 5
            MAX_TEXT_LEN = 10
            dots = "."*3
            rows_len = len(self.data)
            row_dots =  rows_len > SHOW_ROWS * 2 + 1
            cols_len = len(self.columns)
            col_dots = cols_len > SHOW_COLS * 2 + 1
            rows_index = [*range(SHOW_ROWS), *range(rows_len-SHOW_ROWS, rows_len)] if row_dots else range(rows_len)
            dots_insert = lambda item: ((*item[:SHOW_COLS], dots, *item[-SHOW_COLS:]) if col_dots else item)
            columns = ["#", *dots_insert(self.columns)]
            data = [[i, *dots_insert(self.data[i])] for i in rows_index]
            data.insert(SHOW_ROWS, [dots]*len(columns)) if row_dots else None
            text_short = lambda text: f"{text.split('\n')[0][:MAX_TEXT_LEN]}..." if len(text) > MAX_TEXT_LEN else text
            frame = [[text_short("None" if value is None else str(value)) for value in item] for item in [columns] + data]
            row_format = ["{:>" + str(width) + "}" for width in self.max_value_len(frame)]
            set_row = lambda row: "  ".join([row_format[i].format(value) for i, value in enumerate(row)])
            return "\n".join(set_row(row) for row in frame) + f"\n\n[{len(self.data)} rows x {len(self.columns)} columns]"
        else:
            return "Empty dataset."
        
    def __repr__(self) -> str:
        return f"Dataset object ({len(self.data)} rows, {len(self.columns)} columns)"

    def __enter__(self) -> Self:
        return self

    def close(self) -> None:
        del self.data
        del self.extra_data
        del self.columns

    def __exit__(self, exc_type, exc_value, exc_traceback) -> None:
        self.close()


def dataset_to_excel(file_name, *datasets: Dataset, sheet_names=None, formatted: bool=True, stream: bool=False) -> str|BytesIO:
    os.remove(file_name) if os.path.exists(file_name) else None
    sheet_names = to_list(sheet_names)
    wb = Workbook()
    for i, ds in enumerate(datasets):
        ws = wb.active if i==0 else wb.create_sheet()
        title = sheet_names[i] if sheet_names else f'Sheet{i+1}'
        ws.title = title
        if ds.extra_data:
            for extra_data in ds.extra_data:
                ws.append(extra_data)
        extra_sheets = 0
        for data in yield_list(ds.data, XLSX_MAX_ROWS - 1 - len(ds.extra_data)):
            if extra_sheets:
                ws = wb.create_sheet()
                ws.title = f'{title}_ext{extra_sheets}'
                fields_row = 1
            else:
                fields_row = len(ds.extra_data) + 1
            ws.append(ds.columns)
            for row in data:
                ws.append(row)
            if formatted:
                # autofilter header
                ws.auto_filter.ref = ws.dimensions.replace('A1:', f'A{fields_row}:')
                side = Side(border_style='thin', color='000000')
                border = Border(top=side, bottom=side, left=side, right=side)
                for r, row in enumerate(ws[ws.dimensions], start=1):
                    for cell in row:
                        if r <= fields_row:
                            cell.font = Font(bold=True, size=cell.font.size + 1)
                            if r == fields_row: 
                                cell.border = border
                                cell.fill = PatternFill("solid", start_color="D7E4BC")
                        else:
                            cell.border = border
                            if "datetime" == numbers.is_datetime(cell.number_format):
                                cell.number_format = "YYYY-mm-dd HH:MM:SS"
                    # merge extra_data
                    # if r < fields_row:
                    #     ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column)
                ws.freeze_panes = ws[f'A{fields_row+1}']
                # auto_with column
                for col in ws.iter_cols():
                    fieldname_len = len(f"{col[fields_row-1].value}")
                    maxvalue_len = max([len(f"{cell.value}") for i, cell in enumerate(col, start=1) if i > fields_row])
                    length = max(fieldname_len, maxvalue_len)
                    length = (length + (4.23 if fieldname_len >= maxvalue_len - 1 else 1.23) if ws.auto_filter else length)*1.23
                    ws.column_dimensions[col[fields_row-1].column_letter].width = length
            extra_sheets += 1
    if stream:
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        buffer.name = file_name
        wb.close()
        return buffer
    else:
        wb.save(file_name)
        wb.close()
        return file_name

def excel_to_dataset(file_name, sheet_name=None, column_index: int|bool=1, empty_cols:bool=False, as_list=True) -> list|dict:
    data = [] if as_list else dict()
    wb = load_workbook(file_name)
    if sheet_name:
        if isinstance(sheet_name, str):
            sheet_name = to_list(sheet_name)
        elif isinstance(sheet_name, int):
            sheet_name = [sheet_name]
    for sheet in (sheet_name or wb.sheetnames):
        ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[sheet]
        ds = Dataset(ws, column_index=column_index, empty_cols=empty_cols)
        data.append(ds) if isinstance(data, list) else data.update({ws.title: ds})
    wb.close()
    return data
